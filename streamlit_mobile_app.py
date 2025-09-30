# streamlit_mobile_app.py
import io
import time
from typing import List, Tuple
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="RP Logger (Mobile)", page_icon="📱", layout="centered")

# --- Mobile-friendly CSS tweaks ---
st.markdown(
    """
    <style>
    .stButton>button {font-size: 22px; padding: 14px 20px; width: 100%;}
    .stNumberInput > div > div > input {font-size: 20px;}
    .stSelectbox > div > div {font-size: 20px;}
    .stTextInput > div > div > input {font-size: 20px;}
    .big {font-size: 20px;}
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("📱 RP Set Logger — Mobile")
st.caption("Upload your tracker (.xlsx), pick Week/Day/Exercise, log sets quickly.")

DEFAULT_WEEKS = ["Week 1", "Week 2", "Week 3", "Week 4", "Deload"]
DEFAULT_DAYS = ["Push A", "Pull A", "Legs A", "Push B", "Pull B", "Legs B"]

uploaded = st.file_uploader("Upload tracker", type=["xlsx"], label_visibility="collapsed")
if not uploaded:
    st.info("Upload your Excel tracker to start.")
    st.stop()

content = uploaded.read()
wb = load_workbook(io.BytesIO(content))

# Week
weeks = [s for s in DEFAULT_WEEKS if s in wb.sheetnames] or [s for s in wb.sheetnames if s.lower().startswith("week") or s.lower()=="deload"]
week = st.selectbox("Week", weeks, index=0)
ws = wb[week]

# Gather days & exercises
# Days are in Column A, exercises Column B
rows = list(ws.iter_rows(min_row=2, values_only=True))
all_days = [r[0] for r in rows if r and r[0]]
ordered_days = [d for d in DEFAULT_DAYS if d in all_days] + [d for d in all_days if d not in DEFAULT_DAYS]

if not ordered_days:
    ordered_days = DEFAULT_DAYS

# Persist selections
if 'sel_day' not in st.session_state:
    st.session_state.sel_day = ordered_days[0]
if 'ex_idx' not in st.session_state:
    st.session_state.ex_idx = 0

st.header("Day")
sel_day = st.selectbox("Day", ordered_days, index=ordered_days.index(st.session_state.sel_day))
st.session_state.sel_day = sel_day

# Build exercise list for selected day
ex_list = [(i+2, r[1]) for i, r in enumerate(rows) if r and r[0]==sel_day and r[1]]
ex_names = [ex for (_row, ex) in ex_list]

st.header("Exercise")
if ex_names:
    st.write(f"**{sel_day}**: {', '.join(ex_names)}", unsafe_allow_html=True)
    st.session_state.ex_idx = st.number_input("Exercise #", min_value=0, max_value=max(0, len(ex_names)-1), value=st.session_state.ex_idx, step=1)
    row_idx, ex_name = ex_list[st.session_state.ex_idx]
else:
    ex_name = st.text_input("Exercise (custom)")
    row_idx = None

if not ex_name:
    st.stop()

# Read targets
def read_cell(row, col):
    return ws.cell(row=row, column=col).value

if row_idx:
    planned_sets = int(read_cell(row_idx, 4) or 3)
    reps_target = read_cell(row_idx, 5) or ""
    rir_target = read_cell(row_idx, 6) or ""
    st.markdown(f"**Planned Sets:** {planned_sets}  ·  **Reps:** {reps_target}  ·  **RIR:** {rir_target}")
else:
    planned_sets = st.number_input("Planned Sets", 1, 5, value=3)

# Inputs for sets
st.header("Log Sets")
weights = []
reps = []
for i in range(1, 6):
    c1, c2 = st.columns(2)
    with c1:
        w = st.number_input(f"Set {i} — Weight", min_value=0.0, value=0.0, step=2.5, format="%.2f")
    with c2:
        r = st.number_input(f"Set {i} — Reps", min_value=0, value=0, step=1)
    weights.append(None if w==0 else w)
    reps.append(None if r==0 else r)

st.subheader("Top Set (optional)")
colA, colB = st.columns(2)
with colA:
    top_wt = st.number_input("Top Weight", min_value=0.0, value=0.0, step=2.5, format="%.2f")
with colB:
    top_reps = st.number_input("Top Reps", min_value=0, value=0, step=1)
if top_wt == 0: top_wt = None
if top_reps == 0: top_reps = None

# Write function

def write_sets(target_row: int):
    # Sets go to columns G..P, Q..R top set
    for i in range(5):
        if weights[i] is not None:
            ws.cell(row=target_row, column=7 + i*2).value = weights[i]
        if reps[i] is not None:
            ws.cell(row=target_row, column=8 + i*2).value = reps[i]
    if top_wt is not None:
        ws.cell(row=target_row, column=17).value = top_wt
    if top_reps is not None:
        ws.cell(row=target_row, column=18).value = top_reps

# Save area
st.divider()
if st.button("💾 Save to Workbook & Download"):
    # If custom exercise, create new row at bottom
    if row_idx is None:
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1).value = sel_day
        ws.cell(row=new_row, column=2).value = ex_name
        ws.cell(row=new_row, column=4).value = planned_sets
        # set e1RM and total vol formulas
        ws.cell(row=new_row, column=19).value = f"=IFERROR(Q{new_row}*(1+R{new_row}/30),\"\")"
        ws.cell(row=new_row, column=20).value = f"=IFERROR(G{new_row}*H{new_row}+I{new_row}*J{new_row}+K{new_row}*L{new_row}+M{new_row}*N{new_row}+O{new_row}*P{new_row},\"\")"
        write_sets(new_row)
    else:
        write_sets(row_idx)

    bio = io.BytesIO()
    wb.save(bio)
    st.success("Saved! Download below.")
    st.download_button(
        label="⬇️ Download Updated Workbook",
        data=bio.getvalue(),
        file_name="RP_Hypertrophy_Tracker_Updated.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

st.caption("Tip: Add this page to your phone's Home Screen for 1-tap access.")
