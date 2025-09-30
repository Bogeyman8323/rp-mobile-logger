# app.py
import io, os
import streamlit as st
from datetime import datetime
from dateutil import tz
from openpyxl import load_workbook, Workbook
from google_auth import build_auth_url, exchange_code_for_token, get_creds, sign_out
from gdrive import DriveClient, MIMETYPE_XLSX
from workout_logic import recommend

st.set_page_config(page_title="Workout Logger (Mobile)", page_icon="📱", layout="centered")

# ---- Mobile CSS
st.markdown("""
<style>
.stButton>button {font-size:22px;padding:12px 18px;width:100%;}
.stNumberInput input, .stTextInput input, .stSelectbox div {font-size:18px;}
</style>
""", unsafe_allow_html=True)

st.title("📱 Workout Logger — Google Drive")

# ---- AUTH
with st.expander("Authentication", expanded=True):
    creds = get_creds() or exchange_code_for_token()
    if not creds:
        st.write("Sign in with Google to continue.")
        st.link_button("🔐 Sign in with Google", build_auth_url())
        st.stop()
    else:
        c1, c2 = st.columns([3,1])
        c1.success("Signed in")
        if c2.button("Sign out"):
            sign_out(); st.rerun()

drive = DriveClient(get_creds())

# ---- FILE PICKER
st.header("Select your tracker")
st.caption("Least‑privilege: scope is `drive.file`. Pick your `.xlsx` tracker to grant access only to that file.")

manual = st.text_input("Google Drive file ID or share link (fallback)", placeholder="paste here")

params = getattr(st, "query_params", None)
params = params.to_dict() if params else st.experimental_get_query_params()
file_id = params.get("fileId")
if isinstance(file_id, list): file_id = file_id[0]

api_key = st.secrets["google"]["api_key"]
app_id  = st.secrets["google"]["app_id"]
access_token = get_creds().token

picker_html = f"""
<!doctype html>
<html>
<head>
  <meta charset=\"utf-8\"/>
  <script src=\"https://apis.google.com/js/api.js\"></script>
</head>
<body style=\"margin:0;padding:0\">
  <button id=\"open\" style=\"width:100%;padding:12px 18px;font-size:18px;\">Open Google Drive Picker</button>
  <script>
    const token=\"{access_token}\";
    const apiKey=\"{api_key}\";
    const appId=\"{app_id}\";
    document.getElementById('open').addEventListener('click', () => {{
      gapi.load('picker', () => {{
        try {{
          const view = new google.picker.DocsView(google.picker.ViewId.DOCS);
          const origin = window.location.origin;
          const picker = new google.picker.PickerBuilder()
            .addView(view)
            .setOrigin(origin)
            .setOAuthToken(token)
            .setDeveloperKey(apiKey)
            .setAppId(appId)
            .setCallback((data) => {{
              if (data.action === google.picker.Action.PICKED) {{
                const id = data.docs[0].id;
                const base = window.location.origin + window.location.pathname;
                window.top.location.href = base + "?fileId=" + encodeURIComponent(id);
              }}
            }})
            .build();
          picker.setVisible(true);
        }} catch (e) {{
          alert("Picker error: "+ e);
        }}
      }});
    }});
  </script>
</body>
</html>
"""

st.components.v1.html(picker_html, height=60)

if not file_id and manual:
    if "id=" in manual:
        file_id = manual.split("id=")[-1].split("&")[0]
    elif "/d/" in manual:
        file_id = manual.split("/d/")[-1].split("/")[0]
    else:
        file_id = manual.strip()

if not file_id:
    st.info("Use the button above (recommended) or paste a file ID/link, then continue.")
    st.stop()

st.success(f"Using fileId: {file_id}")

# ---- DOWNLOAD EXCEL
try:
    content = drive.download_bytes(file_id)
except Exception as e:
    st.error(f"Download failed: {e}")
    st.stop()

try:
    wb = load_workbook(io.BytesIO(content))
except Exception:
    wb = Workbook()

sheet_name = "Logs"
ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
if ws.max_row == 1 and all(ws.cell(1, c).value is None for c in range(1, 5)):
    ws.append(["Date", "Week", "Day", "Exercise", "Set", "Reps", "Weight", "Notes"])

# ---- WORKOUT FLOW
st.header("Log your sets")
c1, c2, c3 = st.columns(3)
with c1: week = st.text_input("Week", "Week 1")
with c2: day  = st.text_input("Day", "Push A")
with c3: when = st.date_input("Date", datetime.now(tz.gettz("US/Central")).date())

exercise = st.text_input("Exercise", placeholder="e.g., Barbell Bench Press")
if not exercise: st.stop()

rec = recommend(exercise)
st.caption(f"Recommendation: **{rec.sets} sets** of **{rec.reps_low}–{rec.reps_high} reps**, target **{rec.rir} RIR**")

n_sets = st.number_input("Sets to log", min_value=1, max_value=10, value=rec.sets, step=1)

weights, reps, notes = [], [], st.text_input("Notes (optional)")
for i in range(1, n_sets+1):
    s1, s2 = st.columns(2)
    with s1:
        w = st.number_input(f"Set {i} — Weight", min_value=0.0, value=0.0, step=2.5, format="%.2f")
    with s2:
        r = st.number_input(f"Set {i} — Reps", min_value=0, value=max(rec.reps_low, 1), step=1)
    weights.append(w); reps.append(r)

if st.button("💾 Save to Drive"):
    try:
        for i, (w, r) in enumerate(zip(weights, reps), start=1):
            ws.append([when.isoformat(), week, day, exercise, i, r, w, notes])
        out = io.BytesIO()
        wb.save(out); out.seek(0)
        drive.update_bytes(file_id, out.read(), MIMETYPE_XLSX)
        st.success("Saved! Your log was appended to the 'Logs' sheet in your Excel file.")
    except Exception as e:
        st.error(f"Upload failed: {e}")
