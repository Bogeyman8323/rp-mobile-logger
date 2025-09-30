import io
from pathlib import Path
import streamlit as st
from openpyxl import load_workbook
from google_auth import build_auth_url, exchange_code_for_token, get_creds, sign_out
from gdrive_io import DriveClient, MIMETYPE_XLSX

st.set_page_config(page_title="RP Logger (Mobile + Google Drive)", page_icon="📱", layout="centered")

st.title("📱 RP Set Logger — Mobile + Google Drive (Drive API listing + search)")

# --- Auth ---
with st.expander("Authentication", expanded=True):
    creds = get_creds() or exchange_code_for_token()
    if not creds:
        url = build_auth_url()
        st.markdown(f"[🔐 Sign in with Google]({url})")
        st.stop()
    else:
        st.success("Signed in with Google Drive")
        if st.button("Sign out"):
            sign_out()
            st.rerun()

client = DriveClient(get_creds())

st.header("Select your tracker (Drive file search)")
st.caption("This app requests a Drive scope so it can list files. Search, filter by owner, and load more pages as needed.")

# --- Search controls ---
DEFAULT_PAGE_SIZE = 25
search_col, owner_col, size_col = st.columns([3, 2, 1])

with search_col:
    search_term = st.text_input("Filename contains (optional)", placeholder=".xlsx or part of file name")
with owner_col:
    owner_filter = st.text_input("Owner name (optional)", placeholder="owner display name")
with size_col:
    page_size = int(st.number_input("Page size", min_value=5, max_value=200, value=DEFAULT_PAGE_SIZE, step=5))

# Build Drive query (restrict to .xlsx files)
query_base = f"mimeType='{MIMETYPE_XLSX}'"
if search_term:
    sanitized = search_term.replace("'", "\\'")
    query_base += f" and name contains '{sanitized}'"

# Initialize session state for pagination & accumulated results
if 'drive_files' not in st.session_state:
    st.session_state.drive_files = []
if 'drive_next_token' not in st.session_state:
    st.session_state.drive_next_token = None
if 'drive_query' not in st.session_state or st.session_state.drive_query != query_base or st.session_state.get('drive_page_size') != page_size or st.session_state.get('drive_owner_filter') != owner_filter:
    # If query/filters changed since last load, clear accumulated files so search starts fresh
    st.session_state.drive_files = []
    st.session_state.drive_next_token = None
    st.session_state.drive_query = query_base
    st.session_state.drive_page_size = page_size
    st.session_state.drive_owner_filter = owner_filter

def load_next_page():
    try:
        resp = client.list_files(q=st.session_state.drive_query, page_size=st.session_state.drive_page_size, page_token=st.session_state.drive_next_token)
        new_files = resp.get("files", [])
        # Optionally filter by owner display name client-side
        if st.session_state.drive_owner_filter:
            of = st.session_state.drive_owner_filter.strip().lower()
            filtered = []
            for f in new_files:
                owners = f.get("owners", []) or []
                owner_names = [o.get("displayName", "").lower() for o in owners]
                if any(of in n for n in owner_names):
                    filtered.append(f)
            new_files = filtered
        # Append unique by id (avoid duplicates across pages)
        existing_ids = {f["id"] for f in st.session_state.drive_files}
        to_add = [f for f in new_files if f["id"] not in existing_ids]
        st.session_state.drive_files.extend(to_add)
        st.session_state.drive_next_token = resp.get("nextPageToken")
    except Exception as e:
        st.error(f"Failed to list Drive files: {e}")

# First load button (or auto-load if empty)
if not st.session_state.drive_files:
    if st.button("Search Drive"):
        load_next_page()
else:
    st.write(f"Found {len(st.session_state.drive_files)} file(s) (page token present: {bool(st.session_state.drive_next_token)})")
    # Owner filter dropdown derived from loaded files (includes typed owner filter option as well)
    owners = []
    for f in st.session_state.drive_files:
        o = f.get("owners", [{}])[0].get("displayName", "")
        if o and o not in owners:
            owners.append(o)
    owners_display = ["All"] + owners
    selected_owner = st.selectbox("Filter owner (client-side)", owners_display, index=0)
    # Apply selected_owner filter to displayed list (client-side)
    display_files = st.session_state.drive_files
    if selected_owner != "All":
        display_files = [f for f in display_files if (f.get("owners", [{}])[0].get("displayName", "") == selected_owner)]

    # Build labels and allow selection
    labels = [f"{f['name']} — {f.get('owners',[{}])[0].get('displayName','')}" for f in display_files]
    if labels:
        idx = st.selectbox("Choose tracker (from results)", options=list(range(len(display_files))), format_func=lambda i: labels[i], index=0)
        chosen_file = display_files[idx]
        file_id = chosen_file["id"]
        st.success(f"Selected: {chosen_file['name']}")
    else:
        st.info("No files match the current owner filter. Try 'All' or adjust search.")

    # Load more if available
    if st.session_state.drive_next_token:
        if st.button("Load more"):
            load_next_page()

    # Optionally allow clearing results to run a fresh search
    if st.button("New search / Clear results"):
        st.session_state.drive_files = []
        st.session_state.drive_next_token = None
        st.experimental_rerun()

    # --- Open selected file button (keeps the previous app flow) ---
    if 'file_id' in locals() and file_id and st.button("📥 Open from Drive"):
        try:
            content = client.download_file(file_id)
            st.session_state['wb_bytes'] = content
            st.session_state['current_file_id'] = file_id
            st.success("Workbook loaded. Log your sets below and save back to Drive.")
        except Exception as e:
            st.error(f"Download failed: {e}")

st.divider()

# The rest of the workbook editing UI is unchanged; if wb_bytes is present, load workbook and continue
if 'wb_bytes' in st.session_state:
    wb_bytes = st.session_state['wb_bytes']
    wb = load_workbook(io.BytesIO(wb_bytes))

    DEFAULT_WEEKS = ["Week 1", "Week 2", "Week 3", "Week 4", "Deload"]
    DEFAULT_DAYS = ["Push A", "Pull A", "Legs A", "Push B", "Pull B", "Legs B"]

    weeks = [s for s in DEFAULT_WEEKS if s in wb.sheetnames] or [s for s in wb.sheetnames if s.lower().startswith("week") or s.lower()=="deload"]
    week = st.selectbox("Week", weeks, index=0)
    ws = wb[week]

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    all_days = [r[0] for r in rows if r and r[0]]
    ordered_days = [d for d in DEFAULT_DAYS if d in all_days] + [d for d in all_days if d not in DEFAULT_DAYS]
    if not ordered_days:
        ordered_days = DEFAULT_DAYS

    if 'sel_day' not in st.session_state:
        st.session_state.sel_day = ordered_days[0]
    if 'ex_idx' not in st.session_state:
        st.session_state.ex_idx = 0

    sel_day = st.selectbox("Day", ordered_days, index=ordered_days.index(st.session_state.sel_day))
    st.session_state.sel_day = sel_day

    ex_list = [(i+2, r[1]) for i, r in enumerate(rows) if r and r[0]==sel_day and r[1]]
    ex_names = [ex for (_row, ex) in ex_list]

    if ex_names:
        st.write(f"**{sel_day}**: {', '.join(ex_names)}", unsafe_allow_html=True)
        st.session_state.ex_idx = st.number_input("Exercise #", min_value=0, max_value=max(0, len(ex_names)-1), value=st.session_state.ex_idx, step=1)
        row_idx, ex_name = ex_list[st.session_state.ex_idx]
    else:
        ex_name = st.text_input("Exercise (custom)")
        row_idx = None

    if ex_name:
        if row_idx:
            planned_sets = int(ws.cell(row=row_idx, column=4).value or 3)
            reps_target = ws.cell(row=row_idx, column=5).value or ""
            rir_target = ws.cell(row=row_idx, column=6).value or ""
            st.caption(f"Planned Sets: {planned_sets} · Reps: {reps_target} · RIR: {rir_target}")
        else:
            planned_sets = st.number_input("Planned Sets", 1, 5, value=3)

        weights, reps = [], []
        for i in range(1, 6):
            c1, c2 = st.columns(2)
            with c1:
                w = st.number_input(f"Set {i} — Weight", min_value=0.0, value=0.0, step=2.5, format="%.2f")
            with c2:
                r = st.number_input(f"Set {i} — Reps", min_value=0, value=0, step=1)
            weights.append(None if w==0 else w)
            reps.append(None if r==0 else r)

        colA, colB = st.columns(2)
        with colA:
            top_wt = st.number_input("Top Weight", min_value=0.0, value=0.0, step=2.5, format="%.2f")
        with colB:
            top_reps = st.number_input("Top Reps", min_value=0, value=0, step=1)
        if top_wt == 0: top_wt = None
        if top_reps == 0: top_reps = None

        def write_sets(target_row: int):
            for i in range(5):
                if weights[i] is not None:
                    ws.cell(row=target_row, column=7 + i*2).value = weights[i]
                if reps[i] is not None:
                    ws.cell(row=target_row, column=8 + i*2).value = reps[i]
            if top_wt is not None:
                ws.cell(row=target_row, column=17).value = top_wt
            if top_reps is not None:
                ws.cell(row=target_row, column=18).value = top_reps

        if st.button("Write sets to workbook"):
            if row_idx is None:
                new_row = ws.max_row + 1
                ws.cell(row=new_row, column=1).value = sel_day
                ws.cell(row=new_row, column=2).value = ex_name
                ws.cell(row=new_row, column=4).value = planned_sets
                ws.cell(row=new_row, column=19).value = f"=IFERROR(Q{new_row}*(1+R{new_row}/30),\"\")"
                ws.cell(row=new_row, column=20).value = f"=IFERROR(G{new_row}*H{new_row}+I{new_row}*J{new_row}+K{new_row}*L{new_row}+M{new_row}*N{new_row}+O{new_row}*P{new_row},\"\")"
                write_sets(new_row)
            else:
                write_sets(row_idx)
            out = io.BytesIO()
            wb.save(out)
            st.session_state['wb_bytes'] = out.getvalue()
            st.success("Sets written in memory. Use 'Save back to Drive' to upload.")

        st.divider()
        if st.button("☁️ Save back to Drive"):
            try:
                updated = client.update_file_content(st.session_state['current_file_id'], st.session_state['wb_bytes'], MIMETYPE_XLSX)
                st.success("Uploaded successfully to Google Drive.")
            except Exception as e:
                st.error(f"Upload failed: {e}")
else:
    st.info("Use the search controls above to find a spreadsheet in Drive, then choose and open it.")
                st.error(f"Upload failed: {e}")
else:
    st.info("Pick your tracker with the Google Picker to start logging.")
